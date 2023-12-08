from pypdf import PdfReader
import zipfile
import csv
from io import TextIOWrapper
from openpyxl import load_workbook
import pathlib
import download


def test_download_file():
    download.test_download_csv()
    download.test_download_pdf()
    download.test_download_xlsx()


def test_create_zip():
    directory = pathlib.Path("files/")
    with zipfile.ZipFile("files.zip", mode="w") as archive:
        for file_path in directory.iterdir():
            archive.write(file_path, arcname=file_path.name)


def test_read_files():
    if zipfile.is_zipfile("files.zip"):
        with zipfile.ZipFile("files.zip", "r") as archive:
            with archive.open("test_pdf.pdf") as file_z:
                pdf_f = PdfReader(file_z)
                # print(pdf_f.pages)
                assert "manual content" in pdf_f.pages[0].extract_text()

            with archive.open('test_csv.csv') as file_c:
                csv_f = csv.reader(TextIOWrapper(file_c))
                head = next(csv_f)
                assert "Game Number" in head

            with archive.open("test_xlsx.xlsx") as file_x:
                xlsx_f = load_workbook('files/test_xlsx.xlsx')
                sheet = xlsx_f.active
                assert "MyLinks" in xlsx_f.sheetnames
    else:
        print("File is not a zip file")


